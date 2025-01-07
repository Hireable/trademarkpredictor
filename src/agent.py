import os
import logging
from typing import List, Dict, Any
from dotenv import load_dotenv
from langchain.text_splitter import RecursiveCharacterTextSplitter
from sentence_transformers import SentenceTransformer
from unstructured.partition.pdf import partition_pdf
from openpyxl import load_workbook

from src.config import get_config
from src.utils import get_pinecone_index, upsert_chunks, delete_namespace

load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

BATCH_SIZE = 5  # Adjust batch size as needed

class TrademarkCaseAgent:
    def __init__(self):
        self.config = get_config()

        # Initialize Pinecone index
        self.index = get_pinecone_index(self.config["PINECONE_INDEX_NAME"], self.config["EMBEDDING_DIMENSION"])

        # Initialize embedding model
        self.embedding_model = SentenceTransformer(self.config["EMBEDDING_MODEL_NAME"])

        # Initialize text splitter
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=self.config["CHUNK_SIZE"],
            chunk_overlap=self.config["CHUNK_OVERLAP"],
            length_function=len,
        )

    def process_pdf(self, pdf_path: str, namespace: str = None):

        if namespace is None:
            namespace = os.path.splitext(os.path.basename(pdf_path))[0]

        # Delete existing namespace to prevent duplicates if re-processing the same PDF
        delete_namespace(self.index, namespace)

        try:
            elements = partition_pdf(filename=pdf_path)
            extracted_text = "\n\n".join([str(el) for el in elements])

        except Exception as e:
            logger.error(f"Error partitioning or extracting text from {pdf_path}: {e}")
            return

        chunks = self.text_splitter.split_text(extracted_text)
        embeddings = self.embedding_model.encode(chunks)

        # Prepare and upsert chunks in batches
        upsert_chunks_list = []
        for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
            chunk_id = f"{namespace}-{i}"
            metadata = {
                "text": chunk,
                "filename": os.path.basename(pdf_path),
                "chunk_index": i,
                "namespace": namespace,  # Include namespace in metadata
            }
            upsert_chunks_list.append({"id": chunk_id, "values": embedding.tolist(), "metadata": metadata})

            # Upsert in batches
            if len(upsert_chunks_list) >= BATCH_SIZE:
                upsert_chunks(self.index, upsert_chunks_list, namespace)
                upsert_chunks_list = []  # Clear the batch

        # Upsert any remaining chunks
        if upsert_chunks_list:
            upsert_chunks(self.index, upsert_chunks_list, namespace)

        logger.info(f"Processed and stored PDF: {pdf_path} into namespace: {namespace}")


    def load_wipo_mapping(self):

        try:
            workbook = load_workbook(self.config["WIPO_MAPPING_FILE"])
            sheet = workbook.active

            wipo_data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                class_num, goods_description, basic_number = row
                if all((class_num, goods_description, basic_number)):
                    metadata = {
                        "class": str(class_num),
                        "goods_description": goods_description,
                        "basic_number": str(basic_number),
                    }
                    embedding = self.embedding_model.encode(goods_description).tolist()
                    wipo_data.append({"id": str(basic_number), "values": embedding, "metadata": metadata})

            wipo_index = get_pinecone_index(self.config["WIPO_INDEX_NAME"], self.config["EMBEDDING_DIMENSION"])

            # Upsert data to Pinecone in batches
            batch_size = 5  # Adjust as needed
            for i in range(0, len(wipo_data), batch_size):
                batch = wipo_data[i : i + batch_size]
                upsert_response = wipo_index.upsert(vectors=batch)
                logger.info(f"Upserted WIPO mapping batch {i // batch_size +1 } of { (len(wipo_data) + batch_size - 1) // batch_size }: {upsert_response}")  # Log batch progress

            logger.info(f"Upserted a total of {len(wipo_data)} WIPO mappings.")

        except FileNotFoundError:
            logger.error(f"WIPO mapping file not found: {self.config['WIPO_MAPPING_FILE']}")
        except Exception as e:
            logger.error(f"Error loading or upserting WIPO mapping: {e}")